#!/usr/bin/env python3
"""
Generate WBS Excel report from AutoPASS IP 2.0 kanban data.

Usage:
    kanban --format json report wbs | python3 ../autopass-kanban/scripts/wbs_report.py

    # explicit output path
    kanban --format json report wbs | python3 ../autopass-kanban/scripts/wbs_report.py \
        --output delivery/reports/2026-06-07.001.autopass_ip_2.0_wbs_report.xlsx

Default output is delivery/reports/<date>.<nnn>.autopass_ip_2.0_wbs_report.xlsx
where <nnn> is a zero-padded sequence number that auto-increments each run.

The script reads JSON from stdin (produced by `kanban --format json report wbs`)
and writes an xlsx report with:
  - Hierarchical WBS numbering (phase.epic.story) from core-derived rows
  - SUM formulas for story-point totals on epic and phase rows
  - Planned Start Date and Planned End Date columns from markdown metadata
  - Actual Start Date, Actual End Date, and Actual Period from lifecycle fields
  - Estimated hours from core-derived report rows
  - Sprint burndown prognosis sheet
  - Phase summary sheet
  - Legend sheet
"""

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── Colour palette ────────────────────────────────────────────────────────────

COLOUR_TITLE_BG            = "FF0D1F40"
COLOUR_HEADER_BG           = "FF1A3060"
COLOUR_PHASE_BG            = "FF1F3864"
COLOUR_EPIC_BG             = "FF2E5EAA"
COLOUR_EPIC_DONE_BG        = "FF2E6E45"  # dark green matching epic-blue luminance
COLOUR_STORY_BG            = "FFFFFFFF"
COLOUR_STORY_INPROGRESS_BG = "FFE6D0FF"  # soft purple
COLOUR_STORY_DONE_BG       = "FFD0F0D0"  # soft green
COLOUR_WHITE_FG            = "FFFFFFFF"
COLOUR_DARK_FG             = "FF1F1F1F"

# ── Output column layout (A–P, 16 columns) ───────────────────────────────────
COL_WBS                = 1   # A: hierarchical WBS number (1.1.2)
COL_ID                 = 2   # B: ID (phase code / EP-* / US-*)
COL_TITLE              = 3   # C: Title
COL_MILESTONE          = 4   # D: Milestone
COL_PRIORITY           = 5   # E: Priority
COL_STATUS             = 6   # F: Status
COL_POINTS             = 7   # G: Story Points (SUM formula for epic/phase)
COL_HOURS              = 8   # H: Est Hours
COL_PERIOD             = 9   # I: Planned Period
COL_PLANNED_START_DATE = 10  # J: Planned Start Date
COL_PLANNED_END_DATE   = 11  # K: Planned End Date
COL_ACTUAL_PERIOD      = 12  # L: Actual Period
COL_ACTUAL_START_DATE  = 13  # M: Actual Start Date
COL_ACTUAL_END_DATE    = 14  # N: Actual End Date
COL_COMPLETED_IN_SPRINT = 15  # O: Completed In Sprint
COL_NOTES              = 16  # P: Notes
TOTAL_COLS             = 16

WBS_COLUMN_WIDTHS = {
    "A": 10,   # WBS No
    "B": 14,   # ID
    "C": 55,   # Title
    "D": 28,   # Milestone
    "E": 12,   # Priority
    "F": 15,   # Status
    "G": 11,   # Story Pts
    "H": 11,   # Est Hours
    "I": 16,   # Period
    "J": 18,   # Planned Start Date
    "K": 18,   # Planned End Date
    "L": 17,   # Actual Period
    "M": 17,   # Actual Start Date
    "N": 15,   # Actual End Date
    "O": 24,   # Completed In Sprint
    "P": 45,   # Notes
}

DATE_FMT = "YYYY-MM-DD"


# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_colour: str) -> PatternFill:
    if hex_colour == "00000000":
        return PatternFill(fill_type=None)
    return PatternFill(fill_type="solid", fgColor=hex_colour)


def _font(bold: bool = False, colour: str = COLOUR_WHITE_FG, size: int = 10) -> Font:
    return Font(bold=bold, color=colour, size=size)


def _is_dark(hex_colour: str) -> bool:
    """Return True if the colour is dark enough to warrant white text."""
    rgb = hex_colour[-6:]
    try:
        r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
    except ValueError:
        return True
    return (0.299 * r + 0.587 * g + 0.114 * b) < 160


def apply_row_style(ws, row_num: int, level: int, col_count: int = TOTAL_COLS):
    if level == 0:
        bg, fg, bold = COLOUR_TITLE_BG,  COLOUR_WHITE_FG, True
    elif level == 1:
        bg, fg, bold = COLOUR_HEADER_BG, COLOUR_WHITE_FG, True
    elif level == 2:
        bg, fg, bold = COLOUR_PHASE_BG,  COLOUR_WHITE_FG, True
    elif level == 3:
        bg, fg, bold = COLOUR_EPIC_BG,   COLOUR_WHITE_FG, True
    elif level == 4:
        bg, fg, bold = COLOUR_STORY_BG,  COLOUR_DARK_FG,  False
    else:
        bg, fg, bold = "00000000",        COLOUR_DARK_FG,  False

    for col in range(1, col_count + 1):
        cell           = ws.cell(row=row_num, column=col)
        cell.fill      = _fill(bg)
        cell.font      = _font(bold=bold, colour=fg)
        cell.alignment = Alignment(vertical="center", wrap_text=False)


def _set_date_cell(cell, d):
    if d is None:
        return
    cell.value         = d
    cell.number_format = DATE_FMT


# ── Date helpers ──────────────────────────────────────────────────────────────

def _parse_iso_date(ts_str: str | None) -> date | None:
    if not ts_str:
        return None
    try:
        return datetime.fromisoformat(ts_str).date()
    except (ValueError, TypeError):
        try:
            return date.fromisoformat(ts_str[:10])
        except (ValueError, TypeError):
            return None


def _set_optional_date_cell(ws, row_num: int, col: int, value: str | None):
    parsed = _parse_iso_date(value)
    if parsed:
        _set_date_cell(ws.cell(row_num, col), parsed)


# ── WBS sheet ─────────────────────────────────────────────────────────────────

def _write_title_row(ws, row_num: int, title: str, span: int = TOTAL_COLS):
    ws.row_dimensions[row_num].height = 28
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=span)
    c            = ws.cell(row=row_num, column=1, value=title)
    c.font       = Font(bold=True, color=COLOUR_WHITE_FG, size=13)
    c.fill       = _fill(COLOUR_TITLE_BG)
    c.alignment  = Alignment(horizontal="left", vertical="center")


def _write_header_row(ws, row_num: int, headers: list):
    ws.row_dimensions[row_num].height = 20
    for col, h in enumerate(headers, start=1):
        c            = ws.cell(row=row_num, column=col, value=h)
        c.font       = _font(bold=True)
        c.fill       = _fill(COLOUR_HEADER_BG)
        c.alignment  = Alignment(horizontal="center", vertical="center")


def _set_outline_level(ws, row_num: int, level: int):
    ws.row_dimensions[row_num].outlineLevel = level


def _write_group_row(ws, row_num: int, row_data: dict, level: int):
    ws.row_dimensions[row_num].height = 20 if level == 2 else 18
    ws.cell(row_num, COL_WBS,       value=row_data["wbs"])
    ws.cell(row_num, COL_ID,        value=row_data["id"])
    ws.cell(row_num, COL_TITLE,     value=f"   {row_data['title']}")
    ws.cell(row_num, COL_MILESTONE, value=row_data.get("milestone", ""))
    ws.cell(row_num, COL_PERIOD,    value=row_data.get("planned_period"))
    ws.cell(row_num, COL_PRIORITY,  value=row_data.get("priority", ""))
    ws.cell(row_num, COL_STATUS,    value=row_data.get("status", ""))
    ws.cell(row_num, COL_HOURS,     value=row_data.get("est_hours"))
    _set_optional_date_cell(ws, row_num, COL_PLANNED_START_DATE, row_data.get("planned_start_date"))
    _set_optional_date_cell(ws, row_num, COL_PLANNED_END_DATE, row_data.get("planned_end_date"))
    _set_optional_date_cell(ws, row_num, COL_ACTUAL_START_DATE, row_data.get("actual_start_date"))
    _set_optional_date_cell(ws, row_num, COL_ACTUAL_END_DATE, row_data.get("actual_end_date"))
    ws.cell(row_num, COL_ACTUAL_PERIOD, value=row_data.get("actual_period"))
    ws.cell(row_num, COL_COMPLETED_IN_SPRINT, value=row_data.get("completed_in_sprint"))
    ws.cell(row_num, COL_NOTES, value=row_data.get("notes") or None)
    apply_row_style(ws, row_num, level=level)


def _write_story_row(ws, row_num: int, row_data: dict):
    status                        = (row_data.get("status") or "").lower()
    ws.row_dimensions[row_num].height = 17

    ws.cell(row_num, COL_WBS,    value=row_data["wbs"])
    ws.cell(row_num, COL_ID,     value=row_data["id"])
    ws.cell(row_num, COL_TITLE,  value=f"      {row_data['title']}")
    ws.cell(row_num, COL_STATUS, value=row_data.get("status", ""))
    ws.cell(row_num, COL_POINTS, value=row_data.get("points"))
    ws.cell(row_num, COL_HOURS,  value=row_data.get("est_hours"))
    ws.cell(row_num, COL_PERIOD, value=row_data.get("planned_period"))
    _set_optional_date_cell(ws, row_num, COL_PLANNED_START_DATE, row_data.get("planned_start_date"))
    _set_optional_date_cell(ws, row_num, COL_PLANNED_END_DATE, row_data.get("planned_end_date"))
    _set_optional_date_cell(ws, row_num, COL_ACTUAL_START_DATE, row_data.get("actual_start_date"))
    _set_optional_date_cell(ws, row_num, COL_ACTUAL_END_DATE, row_data.get("actual_end_date"))
    ws.cell(row_num, COL_ACTUAL_PERIOD, value=row_data.get("actual_period"))
    ws.cell(row_num, COL_COMPLETED_IN_SPRINT, value=row_data.get("completed_in_sprint"))
    ws.cell(row_num, COL_NOTES, value=row_data.get("notes") or None)

    apply_row_style(ws, row_num, level=4)

    if status == "in progress":
        status_fill = _fill(COLOUR_STORY_INPROGRESS_BG)
        for col in range(1, TOTAL_COLS + 1):
            ws.cell(row=row_num, column=col).fill = status_fill
    elif status in ("done", "dropped"):
        status_fill = _fill(COLOUR_STORY_DONE_BG)
        for col in range(1, TOTAL_COLS + 1):
            ws.cell(row=row_num, column=col).fill = status_fill


def build_wbs_sheet(ws, rows: list, generated_at: str):
    for col_letter, width in WBS_COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    ws.sheet_properties.outlinePr.summaryBelow = False

    # Hide all columns beyond the data range (P onwards) as a single XML span
    _first_hidden = TOTAL_COLS + 1          # 16 → column P
    _hidden_dim   = ws.column_dimensions[get_column_letter(_first_hidden)]
    _hidden_dim.hidden = True
    _hidden_dim.min    = _first_hidden
    _hidden_dim.max    = 16_384             # last Excel column (XFD)

    # Place the cursor on the first data cell when the sheet is opened
    ws.sheet_view.selection[0].activeCell = "A1"
    ws.sheet_view.selection[0].sqref      = "A1"

    report_date = date.fromisoformat(generated_at[:10])
    _write_title_row(ws, 1, f"AutoPASS IP 2.0 – WBS – Report {report_date.strftime('%Y-%m-%d')}")
    _write_header_row(ws, 2, [
        "WBS No", "ID", "Title", "Milestone", "Priority",
        "Status", "Story Pts", "Est Hours", "Planned Period", "Planned Start Date", "Planned End Date",
        "Actual Period", "Actual Start Date", "Actual End Date", "Completed In Sprint", "Notes",
    ])

    row_num_by_wbs: dict[str, int] = {}
    row = 3
    for report_row in rows:
        kind = report_row["kind"]
        row_num_by_wbs[report_row["wbs"]] = row
        if kind == "phase":
            _write_group_row(ws, row, report_row, level=2)
        elif kind == "epic":
            _write_group_row(ws, row, report_row, level=3)
            _set_outline_level(ws, row, 1)
        else:
            _write_story_row(ws, row, report_row)
            _set_outline_level(ws, row, 2)
        row += 1

    for index, report_row in enumerate(rows):
        kind = report_row["kind"]
        row_num = row_num_by_wbs[report_row["wbs"]]
        if kind == "epic":
            child_rows = [
                row_num_by_wbs[row["wbs"]]
                for row in rows[index + 1:]
                if row["kind"] == "story" and row["wbs"].startswith(f"{report_row['wbs']}.")
            ]
            ws.cell(row_num, COL_POINTS).value = (
                f"=SUM(G{min(child_rows)}:G{max(child_rows)})" if child_rows else 0
            )
            if report_row.get("status") == "DONE":
                done_fill = _fill(COLOUR_EPIC_DONE_BG)
                for col in range(1, TOTAL_COLS + 1):
                    ws.cell(row=row_num, column=col).fill = done_fill
        elif kind == "phase":
            refs = [
                f"G{row_num_by_wbs[row['wbs']]}"
                for row in rows[index + 1:]
                if row["kind"] == "epic" and row["wbs"].startswith(f"{report_row['wbs']}.")
            ]
            ws.cell(row_num, COL_POINTS).value = f"=SUM({','.join(refs)})" if refs else 0

    ws.sheet_view.showGridLines = False


# ── Phase Summary sheet ───────────────────────────────────────────────────────

def build_phase_summary_sheet(ws, phases: list):
    for col, width in zip("ABCDEFGHIJ", [10, 55, 12, 30, 8, 9, 12, 12, 13, 13]):
        ws.column_dimensions[col].width = width

    _write_title_row(ws, 1, "AutoPASS IP 2.0 — Phase & Milestone Summary", span=10)
    _write_header_row(ws, 2, [
        "Phase", "Title", "Period", "Milestone", "Epics", "Stories",
        "Pts Total", "Pts Done", "Pts In Progress", "Pts Remaining",
    ])

    totals = {"epics": 0, "stories": 0, "total": 0, "done": 0, "wip": 0, "remaining": 0}
    row    = 3
    for ph_dto in sorted(phases, key=lambda p: p["phase"]):
        ws.row_dimensions[row].height = 18

        data = [
            ph_dto["phase"],
            ph_dto["title"],
            ph_dto["period"],
            ph_dto["milestone"],
            ph_dto["epics"],
            ph_dto["stories"],
            ph_dto["total"],
            ph_dto["done"],
            ph_dto["wip"],
            ph_dto["remaining"],
        ]
        for col, val in enumerate(data, start=1):
            c            = ws.cell(row=row, column=col, value=val)
            c.font       = _font(bold=False, colour=COLOUR_DARK_FG)
            c.alignment  = Alignment(horizontal="left" if col <= 4 else "center", vertical="center")
            if col >= 5:
                c.fill   = _fill("FFE8F0FA")

        totals["epics"]     += ph_dto["epics"]
        totals["stories"]   += ph_dto["stories"]
        totals["total"]     += ph_dto["total"]
        totals["done"]      += ph_dto["done"]
        totals["wip"]       += ph_dto["wip"]
        totals["remaining"] += ph_dto["remaining"]
        row += 1

    ws.row_dimensions[row].height = 20
    for col, val in enumerate(
        ["TOTAL", "", "", "", totals["epics"], totals["stories"],
         totals["total"], totals["done"], totals["wip"], totals["remaining"]],
        start=1,
    ):
        c            = ws.cell(row=row, column=col, value=val)
        c.font       = Font(bold=True, color=COLOUR_WHITE_FG, size=10)
        c.fill       = _fill(COLOUR_PHASE_BG)
        c.alignment  = Alignment(horizontal="left" if col <= 4 else "center", vertical="center")

    ws.sheet_view.showGridLines = False


# ── Sprint Burndown sheet ─────────────────────────────────────────────────────

def build_sprint_burndown_sheet(ws, sprint_rows: list, velocity: dict, forecast: dict, generated_at: str, daily_avg: float):
    for col, width in zip("ABCDEFGH", [28, 13, 13, 14, 14, 14, 16, 18]):
        ws.column_dimensions[col].width = width

    _write_title_row(ws, 1, "AutoPASS IP 2.0 — Sprint Burndown & Prognosis", span=8)

    avg       = velocity.get("avg_points_per_sprint", 0)
    remaining = velocity.get("remaining_points", 0)
    completed = velocity.get("completed_sprint_count", 0)
    completion = forecast.get("completion", {}) if forecast else {}
    throughput = forecast.get("throughput", {}) if forecast else {}
    observed_days = throughput.get("observed_day_count", 0) or 0
    p50 = completion.get("p50_date")
    p80 = completion.get("p80_date")
    p90 = completion.get("p90_date")
    est_text = f"P50 {p50} / P80 {p80} / P90 {p90}" if p80 else "-"

    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 16
    sc            = ws["A2"]
    sc.value      = (
        f"Throughput: {daily_avg:.1f} pts/workday (over {observed_days} observed workdays)  ·  "
        f"Sprint velocity: {avg:.1f} pts/sprint (over {completed} completed sprint{'s' if completed != 1 else ''})  ·  "
        f"Remaining: {remaining} pts  ·  Forecast completion: {est_text}  ·  Generated: {generated_at[:10]}"
    )
    sc.font       = Font(italic=True, color=COLOUR_WHITE_FG, size=9)
    sc.fill       = _fill(COLOUR_HEADER_BG)
    sc.alignment  = Alignment(horizontal="left", vertical="center")

    _write_header_row(ws, 3, [
        "Sprint", "Start", "End", "Planned Pts", "Delivered Pts",
        "Rate (avg)", "Remaining (cum.)", "Status",
    ])

    row                  = 4

    for sprint in sprint_rows:
        status = sprint.get("status") or "planned"
        is_projected = status.startswith("projected")
        if is_projected:
            row_bg = "FFF5F5F5"
        elif status == "closed":
            status_str, row_bg = "completed", "FFEBF5EB"
        elif status == "active":
            status_str, row_bg = "active",    "FFFFF3CD"
        else:
            status_str, row_bg = "planned",   "FFFFFFFF"
        if is_projected:
            status_str = "projected"

        ws.row_dimensions[row].height = 17
        row_data = [
            sprint["name"],
            sprint["start_date"],
            sprint["end_date"],
            sprint.get("planned_points"),
            sprint.get("delivered_points"),
            sprint.get("rate"),
            sprint.get("remaining"),
            status_str,
        ]
        for col, val in enumerate(row_data, start=1):
            c            = ws.cell(row=row, column=col, value=val)
            c.font       = Font(color="FF888888" if is_projected else COLOUR_DARK_FG, italic=is_projected, size=9 if is_projected else 10)
            c.fill       = PatternFill(fill_type="solid", fgColor=row_bg[2:])
            c.alignment  = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
            if col in (4, 5, 6, 7) and val is not None:
                c.number_format = "0"
        row += 1

    ws.sheet_view.showGridLines = False


# ── Legend sheet ──────────────────────────────────────────────────────────────

def build_legend_sheet(ws):
    for col, width in zip("ABC", [22, 50, 20]):
        ws.column_dimensions[col].width = width

    _write_title_row(ws, 1, "AutoPASS IP 2.0 — Legend & Guide", span=3)

    sections = [
        ("Row types", [
            (COLOUR_PHASE_BG,            "Phase",       "Top-level project phase (F1–F5)"),
            (COLOUR_EPIC_BG,             "Epic",        "Epic grouping user stories within a phase"),
            (COLOUR_STORY_BG,            "Story",       "User story (default / not started)"),
        ]),
        ("Story status colours", [
            (COLOUR_STORY_INPROGRESS_BG, "In Progress", "Story currently being developed"),
            (COLOUR_STORY_DONE_BG,       "Done",        "Completed and accepted story"),
        ]),
        ("WBS columns", [
            (None, "WBS No",     "Hierarchical number (phase.epic.story)"),
            (None, "ID",         "Artifact ID — Fn / EP-Fn-* / US-Fn-*"),
            (None, "Title",      "Phase, epic, or story title"),
            (None, "Milestone",  "Delivery milestone (MP1–MP5)"),
            (None, "Priority",   "Critical / High / Medium / Low"),
            (None, "Status",     "Current workflow status"),
            (None, "Story Pts",  "Estimated story points; SUM for epic/phase rows"),
            (None, "Est Hours",  "Estimated hours (throughput-based)"),
            (None, "Planned Period", "Planned quarter or period derived from planned dates when present"),
            (None, "Planned Start Date", "Stored markdown baseline start date; never recalculated from velocity"),
            (None, "Planned End Date",   "Stored markdown baseline end date; blank when no baseline exists"),
            (None, "Actual Period",      "Quarter or period derived from actual lifecycle dates"),
            (None, "Actual Start Date",  "Lifecycle start date from work_started"),
            (None, "Actual End Date",    "Lifecycle completion date from work_done"),
            (None, "Completed In Sprint", "Terminal stories use retained sprint; completed epics use the inclusive sprint date range containing work_done; blank for phases, groups, and unresolved data"),
            (None, "Notes",              "Missing planned baseline or other report remarks"),
        ]),
    ]

    row = 3
    for section_title, entries in sections:
        ws.row_dimensions[row].height = 18
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        hc            = ws.cell(row=row, column=1, value=section_title)
        hc.font       = Font(bold=True, color=COLOUR_WHITE_FG, size=10)
        hc.fill       = _fill(COLOUR_HEADER_BG)
        hc.alignment  = Alignment(horizontal="left", vertical="center")
        row          += 1

        for colour, label, desc in entries:
            ws.row_dimensions[row].height = 17
            fg            = COLOUR_WHITE_FG if colour and _is_dark(colour) else COLOUR_DARK_FG
            swatch        = ws.cell(row=row, column=1, value=label)
            swatch.font   = Font(bold=bool(colour), color=fg, size=10)
            swatch.fill   = _fill(colour) if colour else PatternFill(fill_type=None)
            swatch.alignment = Alignment(horizontal="left", vertical="center")

            desc_cell           = ws.cell(row=row, column=2, value=desc)
            desc_cell.font      = Font(color=COLOUR_DARK_FG, size=10)
            desc_cell.alignment = Alignment(horizontal="left", vertical="center")
            row                += 1

        row += 1  # blank separator between sections

    ws.sheet_view.showGridLines = False


# ── Output path helpers ───────────────────────────────────────────────────────

REPORT_DIR      = Path("delivery/reports")
REPORT_BASENAME = "autopass_ip_2.0_wbs_report.xlsx"


def _next_output_path() -> Path:
    """Return delivery/reports/<today>.<nnn>.autopass_ip_2.0_wbs_report.xlsx.

    Scans the directory for existing files with today's date prefix and picks
    the next sequence number, starting at 001.
    """
    today = date.today().strftime("%Y-%m-%d")
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    existing = sorted(REPORT_DIR.glob(f"{today}.???.{REPORT_BASENAME}"))
    seq      = (int(existing[-1].name.split(".")[1]) + 1) if existing else 1
    return REPORT_DIR / f"{today}.{seq:03d}.{REPORT_BASENAME}"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate WBS xlsx report from kanban JSON data (read from stdin)."
    )
    parser.add_argument(
        "--output",
        default=None,
        help=(
            "Output path for the generated report (.xlsx). "
            f"Defaults to {REPORT_DIR}/<date>.<nnn>.{REPORT_BASENAME}"
        ),
    )
    args = parser.parse_args()

    output_path = Path(args.output) if args.output else _next_output_path()

    try:
        envelope = json.loads(sys.stdin.read())
    except json.JSONDecodeError as e:
        print(f"ERROR: Failed to parse JSON from stdin: {e}", file=sys.stderr)
        sys.exit(1)

    if envelope.get("status") != "ok":
        msg = envelope.get("error", {}).get("message", "unknown error")
        print(f"ERROR: kanban reported an error: {msg}", file=sys.stderr)
        sys.exit(1)

    data         = envelope["data"]
    stories      = data["stories"]
    velocity     = data["velocity"]
    forecast     = data["forecast"]
    generated_at = data["generated_at"]
    wbs_rows     = data.get("wbs_rows")
    phase_rows   = data.get("phase_rows")
    sprint_rows  = data.get("sprint_rows")
    if wbs_rows is None or phase_rows is None or sprint_rows is None:
        print("ERROR: report JSON is missing precomputed rows; rerun with a current kanban binary.", file=sys.stderr)
        sys.exit(1)
    hpp          = data.get("hours_per_point", 0) or 0
    avg_daily    = data.get("daily_avg", 0) or 0
    source       = data.get("throughput_source", "no throughput data")

    wb = openpyxl.Workbook()

    ws_wbs       = wb.active
    ws_wbs.title = "WBS – AutoPASS IP 2.0"
    print("Building WBS sheet …", file=sys.stderr)
    build_wbs_sheet(ws_wbs, wbs_rows, generated_at)

    print("Building Phase Summary sheet …", file=sys.stderr)
    build_phase_summary_sheet(wb.create_sheet("Phase Summary"), phase_rows)

    print("Building Sprint Burndown sheet …", file=sys.stderr)
    build_sprint_burndown_sheet(wb.create_sheet("Sprint Burndown"), sprint_rows, velocity, forecast, generated_at, avg_daily)

    print("Building Legend sheet …", file=sys.stderr)
    build_legend_sheet(wb.create_sheet("Legend & Guide"))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Report saved: {output_path}", file=sys.stderr)
    print(f"  Stories: {len(stories)}", file=sys.stderr)
    print(f"  Sprints: {len([row for row in sprint_rows if not (row.get('status') or '').startswith('projected')])}", file=sys.stderr)
    print(f"  Phases:  {len(phase_rows)}", file=sys.stderr)
    if hpp > 0:
        print(
            f"  Hours/point: {hpp:.1f}h  "
            f"({source}, {avg_daily:.1f} pts/workday)",
            file=sys.stderr,
        )
    completion = forecast.get("completion", {})
    if completion.get("p80_date"):
        print(
            f"  Forecast: P50 {completion.get('p50_date')} / "
            f"P80 {completion.get('p80_date')} / P90 {completion.get('p90_date')} "
            f"({forecast.get('confidence', 'unknown')} confidence)",
            file=sys.stderr,
        )


if __name__ == "__main__":
    main()
