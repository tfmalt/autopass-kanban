#!/usr/bin/env python3

import unittest
from datetime import date
from pathlib import Path
import sys

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
import wbs_report


def workbook_row(
    kind,
    wbs,
    id,
    title,
    status="",
    points=None,
    est_hours=None,
    planned_period=None,
    planned_start_date=None,
    planned_end_date=None,
    actual_period=None,
    actual_start_date=None,
    actual_end_date=None,
    notes="",
):
    return {
        "kind": kind,
        "wbs": wbs,
        "id": id,
        "title": title,
        "milestone": "MP1 - Foundation",
        "priority": "Critical",
        "status": status,
        "points": points,
        "est_hours": est_hours,
        "planned_period": planned_period,
        "planned_start_date": planned_start_date,
        "planned_end_date": planned_end_date,
        "actual_period": actual_period,
        "actual_start_date": actual_start_date,
        "actual_end_date": actual_end_date,
        "notes": notes,
    }


class WbsReportTests(unittest.TestCase):
    def test_wbs_sheet_renders_precomputed_rows_without_rederiving_dates(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            workbook_row(
                "phase",
                "1",
                "F1",
                "Phase 1 - Etablering (Establishment)",
                status="IN PROGRESS",
                points=3,
                planned_period="Q2 2026",
                planned_start_date="2026-06-15",
                planned_end_date="2026-06-19",
                actual_period="Q2-Q3 2026",
                actual_start_date="2026-06-17",
                actual_end_date="2026-07-01",
            ),
            workbook_row(
                "epic",
                "1.1",
                "EP-F1-06",
                "Git-driven kanban and backlog tooling",
                status="IN PROGRESS",
                points=3,
                planned_period="Q2 2026",
                planned_start_date="2026-06-15",
                planned_end_date="2026-06-19",
                actual_period="Q2-Q3 2026",
                actual_start_date="2026-06-17",
                actual_end_date="2026-07-01",
            ),
            workbook_row(
                "story",
                "1.1.1",
                "US-F1-058",
                "Add planned and actual dates",
                status="DONE",
                points=1,
                est_hours=7,
                planned_period="Q2 2026",
                planned_start_date="2026-06-15",
                planned_end_date="2026-06-19",
                actual_period="Q2-Q3 2026",
                actual_start_date="2026-06-17",
                actual_end_date="2026-07-01",
            ),
            workbook_row(
                "story",
                "1.1.2",
                "US-F1-059",
                "Missing planned dates stay visible",
                status="TODO",
                points=2,
                est_hours=14,
                notes="Missing planned baseline: start, end",
            ),
        ]

        wbs_report.build_wbs_sheet(ws, rows, "2026-06-11T10:00:00+02:00")

        headers = [ws.cell(2, col).value for col in range(1, wbs_report.TOTAL_COLS + 1)]
        self.assertIn("Planned Start Date", headers)
        self.assertIn("Planned End Date", headers)
        self.assertIn("Actual Start Date", headers)
        self.assertIn("Actual End Date", headers)
        self.assertIn("Actual Period", headers)

        self.assertEqual(ws.cell(5, wbs_report.COL_PERIOD).value, "Q2 2026")
        self.assertEqual(ws.cell(5, wbs_report.COL_PLANNED_START_DATE).value, date(2026, 6, 15))
        self.assertEqual(ws.cell(5, wbs_report.COL_PLANNED_END_DATE).value, date(2026, 6, 19))
        self.assertEqual(ws.cell(5, wbs_report.COL_ACTUAL_START_DATE).value, date(2026, 6, 17))
        self.assertEqual(ws.cell(5, wbs_report.COL_ACTUAL_END_DATE).value, date(2026, 7, 1))
        self.assertEqual(ws.cell(5, wbs_report.COL_ACTUAL_PERIOD).value, "Q2-Q3 2026")

        self.assertIsNone(ws.cell(6, wbs_report.COL_PERIOD).value)
        self.assertIsNone(ws.cell(6, wbs_report.COL_PLANNED_START_DATE).value)
        self.assertIsNone(ws.cell(6, wbs_report.COL_PLANNED_END_DATE).value)
        self.assertIsNone(ws.cell(6, wbs_report.COL_ACTUAL_PERIOD).value)
        self.assertEqual(
            ws.cell(6, wbs_report.COL_NOTES).value,
            "Missing planned baseline: start, end",
        )

        self.assertEqual(ws.cell(3, wbs_report.COL_POINTS).value, "=SUM(G4)")
        self.assertEqual(ws.cell(4, wbs_report.COL_POINTS).value, "=SUM(G5:G6)")
        self.assertFalse(ws.sheet_properties.outlinePr.summaryBelow)
        self.assertEqual(ws.row_dimensions[3].outlineLevel, 0)
        self.assertEqual(ws.row_dimensions[4].outlineLevel, 1)
        self.assertEqual(ws.row_dimensions[5].outlineLevel, 2)
        self.assertEqual(ws.row_dimensions[6].outlineLevel, 2)

    def test_done_epic_and_dropped_story_use_done_row_highlight(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            workbook_row("phase", "1", "F1", "Phase 1", status="DONE", points=2),
            workbook_row("epic", "1.1", "EP-F1-03", "Dropped work", status="DONE", points=2),
            workbook_row("story", "1.1.1", "US-F1-010", "Dropped story", status="DROPPED", points=2),
        ]

        wbs_report.build_wbs_sheet(ws, rows, "2026-07-02T10:00:00+02:00")

        story_fill = ws.cell(5, wbs_report.COL_STATUS).fill.fgColor.rgb
        epic_fill = ws.cell(4, wbs_report.COL_STATUS).fill.fgColor.rgb
        self.assertEqual(story_fill, wbs_report.COLOUR_STORY_DONE_BG)
        self.assertEqual(epic_fill, wbs_report.COLOUR_EPIC_DONE_BG)

    def test_phase_summary_consumes_precomputed_phase_rows(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        wbs_report.build_phase_summary_sheet(ws, [
            {
                "phase": "F1",
                "title": "Phase 1 - Etablering (Establishment)",
                "period": "Q2 2026",
                "milestone": "MP1 - Foundation",
                "epics": 2,
                "stories": 3,
                "total": 8,
                "done": 5,
                "wip": 1,
                "remaining": 2,
            }
        ])

        self.assertEqual(ws.cell(3, 1).value, "F1")
        self.assertEqual(ws.cell(3, 5).value, 2)
        self.assertEqual(ws.cell(3, 10).value, 2)
        self.assertEqual(ws.cell(4, 1).value, "TOTAL")
        self.assertEqual(ws.cell(4, 7).value, 8)

    def test_sprint_sheet_consumes_precomputed_projection_rows(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            {
                "name": "S000.start",
                "start_date": "2026-06-01",
                "end_date": "2026-06-14",
                "planned_points": 5,
                "delivered_points": 5,
                "rate": 2.7,
                "remaining": 8,
                "status": "closed",
            },
            {
                "name": "S001.projected",
                "start_date": "2026-06-15",
                "end_date": "2026-06-28",
                "planned_points": 27,
                "delivered_points": 8,
                "rate": 2.7,
                "remaining": 0,
                "status": "projected (daily throughput over 3 observed workdays)",
            },
        ]
        velocity = {"avg_points_per_sprint": 5, "remaining_points": 8, "completed_sprint_count": 1}
        forecast = {"throughput": {"observed_day_count": 3}, "completion": {"p80_date": "2026-06-16"}}

        wbs_report.build_sprint_burndown_sheet(ws, rows, velocity, forecast, "2026-06-10T10:00:00+02:00", 2.7)

        self.assertEqual(ws.cell(4, 1).value, "S000.start")
        self.assertEqual(ws.cell(4, 8).value, "completed")
        self.assertEqual(ws.cell(5, 1).value, "S001.projected")
        self.assertEqual(ws.cell(5, 8).value, "projected")
        self.assertEqual(ws.cell(5, 7).value, 0)


if __name__ == "__main__":
    unittest.main()
